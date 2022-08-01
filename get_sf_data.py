# -*- coding: UTF-8 -*-
import datetime
import csv
import os
import openpyxl
import dateparser
from collections import defaultdict

def initial_insert(db, custs):
    fields = ["inst_id", "backend", "org_id"]
    db.insert("sf_data", fields, custs, pk=True, del_table=True)

def get_act_info(sfdb, inst_ids, db):
    query = f"""
    select i.id,
    a.name,
    a.ARR__c,
    csm.Name,
    cse.Name,
    a.Customer_Success_Manager_Role__c as Tier,
    GS_CSM_Meter_Score__c,
    GS_Overall_Score__c,
    Health_Scores_Updated__c,
    a.Account_ID_18_Digits__c,
    a.cs_tier__c,
    a.previous_cs_tier__c,
    a.csm_meter_comments__c
    from edw_tesseract.sbu_ref_sbusfdc.account a
    inner join edw_tesseract.sbu_ref_sbusfdc.installation__c i on a.Account_ID_18_Digits__c = i.Account__c
    left join edw_tesseract.sbu_ref_sbusfdc.user_sbu csm on a.Assigned_CP__c = csm.Id
    left join edw_tesseract.sbu_ref_sbusfdc.user_sbu cse on a.Customer_Success_Engineer__c = cse.Id
    where i.id in ('{"','".join(inst_ids)}')
    order by a.name
    """
    data = [[str(x) for x in sublist] for sublist in sfdb.execute(query)]
    fields = ["inst_id", "account_name", "arr", "csm", "cse", "csm_role", "gsm_score", "gs_overall", "gs_last_update_date",  "account_id", "tier", "previous tier", "csm_comments"]
    db.insert("sf_data", fields, data)

    query = f"""
    select i1.id,
    max(i2.Licenses_Purchased__c),
    a.Account__c
    from edw_tesseract.sbu_ref_sbusfdc.installation__c i1
    left join edw_tesseract.sbu_ref_sbusfdc.installation__c i2 on i1.Account__c = i2.Account__c
    left join edw_tesseract.sbu_ref_sbusfdc.account a on i1.Account__c = a.Account_ID_18_Digits__c
    where i1.id in ('{"','".join(inst_ids)}')
    and i2.Cb_Defense_Org_ID__c is not NULL
    and (i2.Status__c in ('New', 'In-Progress', 'Complete') or i2.status__c is null)
    and i2.Installation_Type__c like '%Subscription%'
    and i2.Cb_Defense_Backend_Instance__c <> 'None'
    --and i2.Normalized_Host_Count__c is not NULL
    group by i1.id, a.Account__c
    """
    data = [[str(x) for x in sublist] for sublist in sfdb.execute(query)]
    fields = ["inst_id", "licenses_purchased", "account__c"]
    db.insert("sf_data", fields, data)

def get_installation_info(sfdb, inst_ids, db):
    query = f"""
    select i.id,
    cast(date_parse(i.createddate, '%Y-%m-%dT%H:%i:%S.000+0000') as date),
    date_diff('day', date_parse(i.createddate, '%Y-%m-%dT%H:%i:%S.000+0000'), cast(i.x50_percent_deployed_date__c as date))
    from edw_tesseract.sbu_ref_sbusfdc.installation__c i
    where i.id in ('{"','".join(inst_ids)}')
    """
    data = sfdb.execute(query)
    fields = ["inst_id", "created_date", "days_to_50perc"]
    db.insert("sf_data", fields, data)

    query = f"""
    select i.id, i.account__c
    from edw_tesseract.sbu_ref_sbusfdc.installation__c i
    left join edw_tesseract.sbu_ref_sbusfdc.account a on i.account__c = a.account_id_18_digits__c
    where i.id in ('{"','".join(inst_ids)}')
    """
    data = sfdb.execute(query)
    lookup = {}
    for i in data:
        lookup[i[1]] = i[0]

    accounts = [i[1] for i in data]
    query = f"""
    select account__c, product_group__c, start_date__c, end_date__c, product_family__c, product__c, product_code__c
    from edw_tesseract.sbu_ref_sbusfdc.bit9_subscriptions__c s
    where s.account__c in ('{"','".join(accounts)}')
    """
    data = sfdb.execute(query)
    products = defaultdict(list)
    for i in data:
        if i[1]:
            products[i[0]].append(i[1])
    for i in products:
        products[i]  = list(set(products[i]))
    for i in list(products):
        inst_id = lookup[i]
        products[inst_id] = ", ".join(sorted(products.pop(i)))
    data = [[i, products[i]] for i in products]
    fields = ["inst_id", "products"]
    db.insert("sf_data", fields, data)

def get_opp_info(sfdb, inst_ids, db):

    def lookup_q(opp_date):
        formatstr = "%Y-%m-%d"
        if isinstance(opp_date, str):
            opp_date = datetime.datetime.strptime(opp_date, formatstr)
        opp_date = datetime.datetime(opp_date.year, opp_date.month, opp_date.day)
        qdict = {
        "2021": {"Q1": ["2020-02-01", "2020-04-30"], "Q2": ["2020-05-01", "2020-07-30"], "Q3": ["2020-07-31", "2020-10-29"], "Q4": ["2020-10-30", "2021-01-28"]},
        "2022": {"Q1": ["2021-01-29", "2021-04-29"], "Q2": ["2021-04-30", "2021-07-29"], "Q3": ["2021-07-30", "2021-10-28"], "Q4": ["2021-10-29", "2022-01-27"]},
        "2023": {"Q1": ["2022-01-28", "2022-04-28"], "Q2": ["2022-04-29", "2022-07-28"], "Q3": ["2022-07-29", "2022-10-27"], "Q4": ["2022-10-28", "2023-01-26"]},
        "2024": {"Q1": ["2023-01-27", "2023-04-27"], "Q2": ["2023-04-28", "2023-07-27"], "Q3": ["2023-07-28", "2023-10-26"], "Q4": ["2023-10-27", "2024-01-25"]},
        "2025": {"Q1": ["2024-01-26", "2024-04-25"], "Q2": ["2024-04-26", "2024-07-25"], "Q3": ["2024-07-26", "2024-10-24"], "Q4": ["2024-10-25", "2025-01-23"]},
        "2026": {"Q1": ["2025-01-24", "2025-04-24"], "Q2": ["2025-04-25", "2025-07-24"], "Q3": ["2025-07-25", "2025-10-23"], "Q4": ["2025-10-24", "2026-01-22"]}
        }
        opp_year = opp_date.year
        for year in qdict:
            for q in qdict[year]:
                if opp_date >= datetime.datetime.strptime(qdict[year][q][0], formatstr) and opp_date <= datetime.datetime.strptime(qdict[year][q][1], formatstr):
                    return f"{year} {q}"
        return "Unknown"

    query = f"""
    select i.id, sum(o.ACV_Amount__c), count(o.Id), array_join(array_agg(o.cb_forecast__c), ','), min(o.CloseDate)
    from edw_tesseract.sbu_ref_sbusfdc.opportunity o
    inner join edw_tesseract.sbu_ref_sbusfdc.installation__c i on o.AccountId = i.Account__c
    where o.CloseDate > CURRENT_DATE
    and o."Type" like '%Renewal%'
    and (position('CBD' IN o.Product_Family__c) > 0 OR position('CBTH' IN o.Product_Family__c) > 0)
    and i.id in ('{"','".join(inst_ids)}')
    group by i.id
    """
    data = sfdb.execute(query)
    for x, row in enumerate(data):
        data[x].append(lookup_q(row[4]))
    fields = ["inst_id", "acv", "opp_ct", "forecast", "renewal_date", "renewal_quar"]
    db.insert("sf_data", fields, data)

def get_case_info(sfdb, inst_ids, db):
    ''' Get number of open cases, cases in last 30d '''
    cases = [
        [["inst_id", "total_cases_30d", "cbc_cases_30d"], "from_iso8601_timestamp(c.createddate) > CURRENT_DATE - interval '30' day"],
        [["inst_id", "open_cases", "open_cbc_cases"], "c.status != 'Closed'"]
        ]
    for x, c in enumerate(cases):
        query = f"""
        select
        i.id,
        count(*) as total_cases,
        SUM(CASE when c.Product_Group__c in (
            'Cb Defense',
            'Cb Defense for VMware',
            'Cb ThreatHunter',
            'Cb ThreatSight',
            'CB Workload') then 1 else 0 end) as cbc_count
        from edw_tesseract.sbu_ref_sbusfdc.installation__c i
        left join edw_tesseract.sbu_ref_sbusfdc.case_sbu c on i.Account__c = c.AccountId
        where i.Id in ('{"','".join(inst_ids)}')
        and {cases[x][1]}
        group by i.Id"""
        data = sfdb.execute(query)
        fields = cases[x][0]
        db.insert("sf_data", fields, data)

def get_cta_info(sfdb, inst_ids, db, cta_type):
    query = f"""
    select a.account_ID_18_Digits__c,
    i.id
    from edw_tesseract.sbu_ref_sbusfdc.installation__c i
    left join edw_tesseract.sbu_ref_sbusfdc.account a on i.Account__c = a.Account_ID_18_Digits__c
    where i.id in ('{"','".join(inst_ids)}')
    """
    inst_ids_acct = sfdb.execute(query, dict=True)
    accts = list(inst_ids_acct)
    query = f"""
    select account_id,
    max(closed_date),
    case when status in ('New','Work In Progress') then 'Open' else 'Closed' end
    from edw_tesseract.sbu_ref_sbusfdc.gsctadataset
    where reason like '{cta_type}'
    and account_id in ('{"','".join(accts)}')
    and status not in ('Closed No Action', 'Closed Unsuccessful', 'Closed Invalid')
    group by account_id, status, closed_date
    """
    rows = []
    data = sfdb.execute(query)
    for row in data:
        for i in inst_ids_acct[row[0]]:
            rows.append([i] + row[1:])
    data_avail = [i[0] for i in rows]
    for inst_id in inst_ids:
        if inst_id not in data_avail:
            rows.append([inst_id, "None", "None"])
    if cta_type == "Product Usage Analytics":
        fields = ["inst_id", "last_cua", "cua_status"]
    elif cta_type == "CSA Whiteboarding":
        fields = ["inst_id", "last_wb"]
        for x, row in enumerate(rows):
              rows[x] = row[:-1]
    elif cta_type == "Tech Assessment":
        fields = ["inst_id", "last_ta"]
        for x, row in enumerate(rows):
            rows[x] = row[:-1]
    db.insert("sf_data", fields, rows)

def get_activity(db):
    xlsx_files = [i for i in os.listdir() if i.endswith(".xlsx") and i.startswith("Distinct")]
    data = []
    for f in xlsx_files:
        wb = openpyxl.load_workbook(f, data_only=True)
        s = wb["Mda Sheet"]
        for x, i in enumerate(s.rows):
            account = s.cell(row=x+1, column=1).value
            act_date = s.cell(row=x+1, column=6).value
            act_date = dateparser.parse(act_date)
            if not act_date:
                continue
            act_date = datetime.datetime.strftime(act_date, "%Y-%m-%d")
            data.append([account, act_date])
    fields = ["account", "activity_date"]
    db.insert("cse_activity", fields, data, pk=False, del_table=True)

if __name__ == "__main__":
    import db_connections
    sfdb = db_connections.tesseract_connection()
    db = db_connections.sqlite_db("cua.db")
    with open("report_setup_tess.sql", "r") as f:
        query = f.read()
    custs = sfdb.execute(query)
    inst_ids = [i[0] for i in custs]
    initial_insert(db, custs)
    get_act_info(sfdb, inst_ids, db)
    get_activity(db)
    get_installation_info(sfdb, inst_ids, db)
    get_opp_info(sfdb, inst_ids, db)
    get_case_info(sfdb, inst_ids, db)
    get_cta_info(sfdb, inst_ids, db, "Product Usage Analytics")
    get_cta_info(sfdb, inst_ids, db, "Tech Assessment")
    get_cta_info(sfdb, inst_ids, db, "CSA Whiteboarding")
